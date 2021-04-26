import numpy as np
import node2vec
import networkx as nx
import math
import time
import random
import openpyxl
from cvxopt import matrix, solvers
def access(route:list, let:list, edges:list, edges_sta:list, zeta:float):
    J = 0
    M = 0
    start = route[-1]
    for i in range(len(route)-2,-1,-1):
        end = start
        start = route[i]
        weight = edges[start][end]
        sta = edges_sta[start][end]
        M += (weight**2 + sta**2 + 2*weight*J)
        J += weight
        # print("J = " + str(J) + " ,i =" + str(i))
        # print("M = " + str(M) + " ,i =" + str(i))
    C = J + zeta * math.sqrt(M-J**2)
    J1 = 0
    M1 = 0
    start = let[-1]
    for i in range(len(let)-2,-1,-1):
        end = start
        start = let[i]
        weight = edges[start][end]
        sta = edges_sta[start][end]
        M1 += (weight**2 + sta**2 + 2*weight*J1)
        J1 += weight
    C1 = J1 + zeta * math.sqrt(M1-J1**2)
    # print("C = " + str(C))
    # print("C1 = "+ str(C1))
    return (C/C1 - 1)

def main():
#     nodes : 4807  Link:  11140
    zeta = 1.64
    nodes = [ 1+x for x in range(1020)]
    edges = []
    data = openpyxl.load_workbook("/Users/pqh/Desktop/route/Barcelona/Barcelona.xlsx")
    worksheet = data.worksheets[0]
    for i in range(2522):
        inn1 = []
        s = worksheet.cell(i+1,1).value
        e = worksheet.cell(i+1,2).value
        w = worksheet.cell(i+1,3).value
        inn1.append(s)
        inn1.append(e)
        inn1.append(w)
        edges.append(inn1)
    edges_wei = np.load(file = "/Users/pqh/Desktop/route/Barcelona/Weight.npy")
    edges_sta = np.load(file = "/Users/pqh/Desktop/route/Barcelona/Standard_t10.6.npy")
    node_to_node = np.load(file = "/Users/pqh/Desktop/route/Barcelona/Node_to_node.npy",allow_pickle=True)
    re_time = []
    re_opt = []
    G = nx.Graph()
    G.add_nodes_from(nodes)
    G.add_weighted_edges_from(edges)

    d = 204
    Tmax = 20
    d_arr = [102,204,510,928]
    end_arr = [1,36,77,85,90,108,209,245,298,366,432,444,574,630,742,789,863,877,934,999]
    start_arr = [3,8,9,11,16,19,22,24,29,33,45,53,58,62,64,69,70,79,98,101,102,204,206,213,215,218,232,
                 244,250,259,261,273,286,296,300,308,313,314,322,338,347,354,355,372,387,400,409,416,438,
                 458,469,477,489,502,506,509,518,526,537,545,555,566,587,593,601,607,612,623,638,642,654,
                 660,675,688,694,701,706,719,720,733,757,762,793,800,809,818,832,843,869,888,899,903,944,
                 969,978,982,991,1003,1015,1018]
    phi = np.load(file = "/Users/pqh/Desktop/route/Barcelona/Barcelona_"+ str(d) + "_phi.npy", allow_pickle = True)
    for end in end_arr:
        p = nx.shortest_path(G,target = end, weight='weight', method='dijkstra')
        next_node = [0 for x in range(1021)]
        len1 = 0
        real_node = []
        for i in range(1021):
            if i in p and i != end:
                real_node.append(i)
                next_node[i] = p.get(i)[1]
                len1 += 1
        print("len1 = " + str(len1))
        t = 0
        c_arr = []
        time_arr = []
        start_time = time.clock()
        while( t < Tmax):

            GG = np.zeros((len1,d+1))
            hh = np.zeros(len1)
            A = np.zeros((len1,d+1))
            b = np.zeros(len1)
            b1 = np.zeros(len1)
            J = [0 for x in range(1021)]
            M = [0 for x in range(1021)]
            for i in range(len1):
                start_node = real_node[i]
                end_node = next_node[start_node]
                wei = edges_wei[start_node][end_node]
                b[i] = wei
                for j in range(d+1):
                    GG[i][j] = -phi[start_node][j]
                    if end_node != end:
                        A[i][j] = phi[start_node][j] - phi[end_node][j]
                    else:
                        A[i][j] = phi[start_node][j]
            # Q = 2 * A.T * A  ; q = A * b; G = - P
            Q1 = matrix(2*np.dot(A.T,A))
            q1 = matrix(-2*np.dot(A.T,b))
            GG1 = matrix(GG)
            hh1 = matrix(hh)
            w_j1 = solvers.qp(Q1,q1,GG1,hh1).get('x')
            w_j = np.array(w_j1).T
            #----------------
            # w_j = np.dot(np.dot(np.linalg.inv(np.dot(A.T,A)),A.T),b)
            for k in real_node:
                J[k] = np.dot(w_j,phi[k])
            for i in range(len1):
                start_node = real_node[i]
                hh[i] = -1* J[start_node]**2
                end_node = next_node[start_node]
                wei = edges_wei[start_node][end_node]
                sta = edges_sta[start_node][end_node]
                b1[i] =wei**2 + sta**2 + 2 * wei * J[end_node]
            # Q = 2 * A.T * A ; q = A * b; G = - P ; h = J
            q1 = matrix(-2*np.dot(A.T,b1))
            hh1 = matrix(hh)
            w_m1 = solvers.qp(Q1,q1,GG1,hh1).get('x')
            w_m = np.array(w_m1).T
            # ------------------
            # w_m = np.dot(np.dot(np.linalg.inv(np.dot(A.T,A)),A.T),b1)
            for k in real_node:
                M[k] = np.dot(w_m,phi[k])
            c = 0
            li = list(real_node)
            random.shuffle(li)
            for l in li:
                neighbor = node_to_node[l]
                l_next = next_node[l]
                l_min = edges_wei[l][l_next] + J[l_next] + zeta * math.sqrt(edges_sta[l][l_next]**2 + M[l_next] - J[l_next]**2)
                for m in neighbor:
                    n = m
                    flag = 0
                    f = 0
                    while(f < 100):
                        if(n == end):
                            flag = 1
                            break
                        if(n == l):
                            break
                        n = next_node[n]
                        f+=1
                    if(flag == 0):
                        continue
                    wei = edges_wei[l][m]
                    sta = edges_sta[l][m]
                    l_cur = wei + J[m] + zeta * math.sqrt(sta**2 + M[m] - J[m]**2)
                    if(l_min > l_cur + 0.00001):
                        l_min = l_cur
                        l_next = m
                if(next_node[l] != l_next):
                    next_node[l] = l_next
                    c+=1
            print("c = " + str(c))
            print("t = " + str(t))
            c_arr.append(c)
            t+=1
            # end_time = time.clock()
            # time_arr.append(end_time - start_time)
            # print("time = " + str(time_arr[-1]))
        end_time = time.clock()
        print("time = "+ str(end_time-start_time))
        re_time.append(end_time-start_time)
        print(c_arr)
        OD = []
        OD1 = []
        result = []
        for k in start_arr:
            m = k
            inn = []
            while(m != end):
                inn.append(m)
                m = next_node[m]
            inn.append(end)
            OD.append(inn)
            OD1.append(p.get(k))
        for i in range(len(OD)):
            a = access(OD[i],OD1[i],list(edges_wei),list(edges_sta),zeta)
            result.append(a)
        print(result)
        print(np.mean(np.array(result)))
        re_opt.append(np.mean(np.array(result)))
    print(re_time)
    print(re_opt)
    print(np.mean(re_opt))
if __name__ == '__main__':
    main()
